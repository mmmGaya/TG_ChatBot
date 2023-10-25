import openpyxl

group_number = input("Введите номер группы: ")
excel_file_path = 'data/schedule/25.10.2023.xlsx'
workbook = openpyxl.load_workbook(excel_file_path)

# пары по номеру группы
def get_schedule(group_number):
    schedule = []
    for sheet_name in workbook.sheetnames:
        if sheet_name.startswith('Пара'):
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[3] == group_number:  # Предполагаем, что номер группы находится в 4-м столбце (индекс 3)
                    schedule.append({
                        'Пара': sheet_name,
                        'Аудитория': row[1],  # Предполагаем, что номер аудитории во 2-м столбце (индекс 1)
                        'Преподаватель': row[2]  # Предполагаем, что имя преподавателя в 3-м столбце (индекс 2)
                    })
    return schedule

schedule = get_schedule(group_number)

# вывод
if schedule:
    print(f"Расписание для группы {group_number}:")
    for item in schedule:
        print(f"{item['Пара']}: Аудитория {item['Аудитория']}, Преподаватель: {item['Преподаватель']}")
else:
    print(f"Расписание для группы {group_number} не найдено.")
